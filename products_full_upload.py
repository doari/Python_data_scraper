import os
import re
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import pandas as pd
from bs4 import BeautifulSoup
from PIL import Image
from io import BytesIO
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ChromeDriver 경로 설정
chrome_driver_path = 'C:/Users/#/#/chromedriver-win64/chromedriver.exe'

# 크롬 브라우저 옵션 설정
chrome_options = Options()
# chrome_options.add_argument("--headless")  # 브라우저 창을 표시하지 않음
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

# 웹 드라이버 초기화
service = Service(chrome_driver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)

# 웹 페이지 열기
url = 'https://www.#.com/ko-kr/#-%EC%97%AC%EC%84%B1/%EB%B2%A8%ED%8A%B8/?nav=A009-VIEW-ALL'
driver.get(url)

# 쿠키 허용 팝업 처리
try:
    wait = WebDriverWait(driver, 10)
    cookie_button = wait.until(EC.element_to_be_clickable(
        (By.ID, 'onetrust-accept-btn-handler')))  # 쿠키 허용 버튼이 클릭 가능할 때까지 대기
    cookie_button.click()
    print("쿠키 허용 버튼 클릭")
except Exception as e:
    print(f"쿠키 허용 버튼 클릭 실패: {e}")

# 페이지 스크롤
scroll_pause_time = 2  # 스크롤 후 대기 시간
last_height = driver.execute_script("return document.body.scrollHeight")

while True:
    driver.execute_script(
        "window.scrollTo(0, document.body.scrollHeight);")  # 페이지 끝까지 스크롤
    time.sleep(scroll_pause_time)
    new_height = driver.execute_script("return document.body.scrollHeight")
    if new_height == last_height:  # 더 이상 스크롤할 내용이 없을 경우 루프 종료
        break
    last_height = new_height

# 스크롤 완료 후 페이지 소스 가져오기
html = driver.page_source
driver.quit()

# BeautifulSoup으로 HTML 파싱
soup = BeautifulSoup(html, 'html.parser')

# 이미지 저장 폴더 생성
os.makedirs('images', exist_ok=True)

# 제품 정보 추출
product_data = []
product_wrappers = soup.find_all('div', class_='m-product-listing__wrapper')


def extract_reference_and_code_from_url(url):
    """
    이미지 URL에서 제품 레퍼런스와 코드를 추출하는 함수
    """
    match = re.search(
        r'/([A-Z0-9]+(?:\.[A-Z0-9]+)*)_([A-Z0-9_]+)\.(jpg|png)', url, re.IGNORECASE)
    if match:
        reference = match.group(1)
        code = f"{match.group(1)}_{match.group(2)}"
        return reference, code
    return 'Unknown', 'Unknown'


def download_image(img_url, save_dir, filename):
    """
    이미지 URL에서 이미지를 다운로드하고 지정된 파일명으로 저장하는 함수
    """
    reference, code = extract_reference_and_code_from_url(img_url)
    if reference == 'Unknown' or code == 'Unknown':  # 레퍼런스나 코드가 없을 경우 에러 메시지 출력
        print(f'Error: Unable to extract reference and code from {img_url}')
        return

    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(img_url, headers=headers)
        response.raise_for_status()
        image = Image.open(BytesIO(response.content))

        # 상품번호를 파일명으로 하여 이미지 저장
        file_path = os.path.join(save_dir, f'{filename}.jpg')
        image.save(file_path)
        print(f'Image saved as {file_path}')
    except Exception as e:
        print(f'Failed to download or save image from {img_url}: {e}')


# 제품 정보 추출 및 이미지 다운로드
for idx, wrapper in enumerate(product_wrappers, start=1):
    product_meta = wrapper.find('div', class_='m-product-listing__meta')
    if not product_meta:
        continue

    # 제품명 추출
    title = product_meta.find(
        'div', class_='m-product-listing__meta-title').get_text(strip=True)
    # 금액 정보 추출
    price_tag = product_meta.find('strong', class_='f-body--em')
    price = price_tag.get_text(
        strip=True) if price_tag else 'No price information'
    # 색상 정보 추출
    color_tag = product_meta.find('span', class_='a11y')
    color = color_tag.get_text(strip=True).lstrip(
        ';') if color_tag else 'No color information'

    item_name = title
    reference = 'Unknown'
    code = 'Unknown'
    image_url = None

    # 이미지 URL 추출 및 다운로드
    image_divs = wrapper.find_all('div', class_='m-product-listing__img-img')
    for image_div in image_divs:
        img_tag = image_div.find('img')
        if img_tag:
            image_url = img_tag.get('src') or img_tag.get('data-lazy-src')
            if image_url:
                image_url = re.sub(r'\?.*$', '', image_url)  # URL에서 쿼리스트링 제거
                reference, code = extract_reference_and_code_from_url(
                    image_url)
                download_image(image_url, 'images', idx)  # 이미지 다운로드
                break

    # 제품 데이터를 리스트에 추가
    product_data.append({
        '상품번호': idx,
        '상품명': item_name,
        '금액': price,
        '모델명(레퍼런스)': reference,
        '옵션(색상)': color
    })

# 데이터 엑셀 파일로 저장
df = pd.DataFrame(product_data, columns=[
                  '상품번호', '상품명', '금액', '모델명(레퍼런스)', '옵션(색상)'])
df.to_excel('#.xlsx', index=False)

# 엑셀 파일 스타일 조정
wb = load_workbook('#_products.xlsx')
ws = wb.active

# 헤더 스타일 설정
header_fill = PatternFill(fill_type="solid")
header_font = Font(size=20, color='0077ff')  # 폰트 크기 20, 색상 파란색

# 1행의 셀 스타일 적용
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 열 너비 조정
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 20  # C열: 금액
ws.column_dimensions['D'].width = 40  # D열: 모델명(레퍼런스)
ws.column_dimensions['E'].width = 20  # E열: 옵션(색상)

# A열 가운데 정렬
for cell in ws['A']:
    cell.alignment = Alignment(horizontal='center')

# B열 가운데 정렬
for cell in ws['B']:
    cell.alignment = Alignment(horizontal='center')

# C열(금액) 가운데 정렬
for cell in ws['C']:
    cell.alignment = Alignment(horizontal='center')

# D열(모델명) 왼쪽 정렬
for cell in ws['D']:
    cell.alignment = Alignment(horizontal='left')

# 엑셀 파일 저장
wb.save('#_products.xlsx')

print('데이터와 이미지가 성공적으로 저장되었습니다.')
